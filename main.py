import  numpy as np
from sympy import *
from openpyxl import Workbook
import datetime
import matplotlib.pyplot as plt

# решение в целом
class Solution(object):
    def __init__(self, h, sigma, L, dt, Nt):
        self.h = h
        self.L = L
        self.sigma = sigma

        self.dt = dt
        self.Nt = Nt + 1
        self.N = int(self.L / self.h) + 1

    def save(self, text=str(datetime.datetime.now())):
        try:
            wb = Workbook()
            ws = wb.active
            for i in range(Nt + 1):
                ws.cell(column=1, row=i+2).value = i*self.dt

            for i in range(int(self.L/self.h) + 1):
                ws.cell(column=i + 2, row=1).value = i * self.h

            for i in range(u.shape[0]):
                row = i + 2
                for j in range(u.shape[1]):
                    col = j + 2
                    ws.cell(column=col, row=row).value = u[i][j]
            wb.save(f'..\\{text}.xlsx')
        finally:
            pass
    def plotting(self, text=str("figure 1")):
        step = 50
        step = int(self.Nt/step)

        plt.figure(f'{text} {self.h} {self.sigma} {self.L}  {self.dt} {self.Nt} space')
        for n in range(0, step):
            plt.plot(np.arange(0, self.L +  self.h,  self.h), u[int(Nt/step*n)][:], label='u({},x)'.format(n))
        plt.legend()
        plt.grid()
        plt.show()

        step = 50
        step = int(self.Nt / step)







    def __del__(self):
        del self.h
        del self.sigma
        del self.N
        del self.L
        del self.dt
        del self.Nt


# прогонка
class TridiagonalMatrixAlgorithm(Solution):

    def __init__(self, h, sigma, L, dt, Nt):
        super().__init__(h, sigma, L, dt, Nt)

    def alpha(self):
        if j == 1:
            return 0
        else:
            return -a/(b + c*alfa[j-1])

    def beta(self):
        if j == 1:
            return 1
        else:
            return (-ksi[n][j] - c*beta[j - 1])/(b + c*alfa[j - 1])

    def ksi(self):
            return u[n][j]

    def __del__(self):
        super().__del__()


# явная рахностная схема
class ExplicitDiffScheme(Solution):

    def __init__(self, h, sigma, L, dt, Nt):
        super().__init__(h, sigma, L, dt, Nt)

    def stability(self):
        return 1/(1 + 2 * self.sigma * self.dt/(self.h * self.h))

    def low_boundary(self):
        return 0

    def upper_boundary(self):
        return 20

    def solution(self):
        global u, n, j
        n = 0
        u = np.empty((self.Nt, self.N))

        u[0][:] = 10

        for j in range(0, self.Nt - 1):
            pass

        while n < self.Nt - 1:
            for j in range(1, self.N - 1):
                u[n + 1][j] = u[n][j] + self.sigma * self.dt / (self.h * self.h) * (u[n][j + 1] - 2 * u[n][j] + u[n][j - 1])

            u[n + 1][0] = self.low_boundary()
            u[n + 1][self.N - 1] = self.upper_boundary()
            n = n + 1

        try:
            self.save(f"explicit_diff_scheme_{self.h}_{self.sigma}_{self.L}_{self.dt}_{self.Nt}")
        finally:
            pass

        try:
            self.plotting(("explicit_diff_scheme_").replace("_", " "))
        finally:
            pass

        print('\n\nЯвная разностная схема\n\n')
        print(f'\n\nУсловие устойчивости:\t{self.stability()}')
        print(f'\n\n{u}')

    def __del__(self):
        super().__del__()


# кранка николсона рахностная схема
class KrankNicholsonScheme(TridiagonalMatrixAlgorithm):

    def __init__(self, h, sigma, L, dt, Nt):
        super().__init__(h, sigma, L, dt, Nt)

    def stability(self):
        return (1 - 2 * self.sigma * self.dt / (self.h * self.h)) / (1 + 2 * self.sigma * self.dt / (self.h * self.h))

    def low_boundary(self):
        return 0

    def upper_boundary(self):
        return 20

    def solution(self):
        global u, n, j, a, b, c, alfa, betta, ksi
        u = np.zeros((self.Nt, self.N))
        alfa = np.zeros((self.N))
        betta = np.zeros((self.N))

        n = 0

        a = -self.sigma * self.dt / (2*self.h * self.h)
        b = 1 - 2 * a
        c = a
        ksi = 0

        u[0][:] = 10

        for j in range(0, self.Nt - 1):
            pass #не используется

        while n < self.Nt - 1:

            alfa[0] = 0
            betta[0] = 0

            for j in range(1, self.N - 1):
                ksi = u[n][j] - a * (u[n][j + 1] - 2*u[n][j] + u[n][j - 1])
                alfa[j] = -a / (b + c * alfa[j - 1])
                betta[j] = (ksi - c * betta[j - 1]) / (b + c * alfa[j - 1])

            u[n + 1][self.N - 1] = self.upper_boundary()

            for j in range(self.N - 2, 0, -1):
                u[n + 1][j] = alfa[j] * u[n + 1][j + 1] + betta[j]
            n = n + 1

        try:
            self.save(f"krank_nicholson_scheme_{self.h}_{self.sigma}_{self.L}_{self.dt}_{self.Nt}")
        finally:
            pass

        try:
            self.plotting(("krank_nicholson_scheme_").replace("_", " "))
        finally:
            pass

        print('\n\nСхема Кранка-Николсона\n\n')
        print(f'\n\n{self.stability()}')
        print(f'\n\n{u}')

    def __del__(self):
        super().__del__()


class ImplicitDiffScheme(TridiagonalMatrixAlgorithm):
    # неявная рахностная схема
    def __init__(self, h, sigma, L, dt, Nt):
            super().__init__(h, sigma, L, dt, Nt)

    def stability(self):
            return (1 - 2 * self.sigma * self.dt / (self.h * self.h)) / (1 + 2 * self.sigma * self.dt / (self.h * self.h))

    def low_boundary(self):
            return 0

    def upper_boundary(self):
            return 20

    def solution(self):
        global u

        u = np.zeros((self.Nt, self.N))
        alfa = np.zeros((self.N))
        betta = np.zeros((self.N))

        n = 0

        a = -sigma * self.dt / (self.h * self.h)
        b = 1 - 2 * a
        c = a
        ksi = 0

        u[0][:] = 10

        for j in range(0, self.Nt - 1):
            pass

        while n < self.Nt - 1:
            alfa[0] = 0
            betta[0] = 0

            for j in range(1, self.N - 1):
                ksi = u[n][j]
                alfa[j] = -a / (b + c * alfa[j - 1])
                betta[j] = (ksi - c * betta[j - 1]) / (b + c * alfa[j - 1])

            u[n + 1][self.N - 1] = self.upper_boundary()

            for j in range(self.N - 2, 0, -1):
                u[n + 1][j] = alfa[j] * u[n + 1][j + 1] + betta[j]
            n = n + 1

        try:
            self.save(f"implicit_diff_scheme_{self.h}_{self.sigma}_{self.L}_{self.dt}_{self.Nt}")
        finally:
            pass

        try:
            self.plotting(("implicit_diff_scheme_").replace("_", " "))
        finally:
            pass

        print('\n\nНеявная разностная схема\n\n')
        print(f'\n\n{self.stability()}')
        print(f'\n\n{u}')

    def __del__(self):
        super().__del__()


class SaulevaDiffScheme(Solution):

    def __init__(self, h, sigma, L, dt, Nt):
        super().__init__(h, sigma, L, dt, Nt)

    def stability(self):
        return 1/(1 + 2 * self.sigma * self.dt/(self.h * self.h))

    def low_boundary(self):
        return 0

    def upper_boundary(self):
        return 20

    def solution(self):
        global u

        a = -sigma * self.dt / (self.h * self.h)
        n = 0
        u = np.empty((self.Nt, self.N))

        u[0][:] = 10

        for j in range(0, self.Nt - 1):
            pass

        while n < self.Nt - 2:
            u[n + 1][0] = self.low_boundary()
            for j in range(1, self.N - 1):
                u[n + 1][j] = (u[n][j] - a * (u[n][j + 1] - u[n][j] + u[n + 1][j - 1]))/(1 - a)

            u[n + 1][self.N - 1] = self.upper_boundary()
            u[n + 2][self.N - 1] = self.upper_boundary()

            for j in range(self.N - 2, 1, -1):
                u[n + 2][j] = (u[n + 1][j] - a * (u[n + 2][j + 1] - u[n + 1][j] + u[n + 1][j - 1]))/(1 - a)

            u[n + 2][0] = self.low_boundary()
            n = n + 2
        try:
            self.save(f"sauleva_diff_scheme_{self.h}_{self.sigma}_{self.L}_{self.dt}_{self.Nt}")
        finally:
            pass
        try:
            self.plotting(("sauleva_diff_scheme_").replace("_", " "))
        finally:
            pass

        print('\n\nСаульева разностная схема\n\n')
        print(f'\n\n{self.stability()}')
        print(f'\n\n{u}')

    def __del__(self):
        super().__del__()


if __name__ == '__main__':
    sigma = 1.8
    h = 0.05
    L = 25
    dt = 0.0005
    Nt = 2000
    # sigma = float(input("input [sigma] value: ").replace(",", "."))
    # h = float(input("input [h] value: ").replace(",", "."))
    # L = float(input("input [L] value: ").replace(",", "."))
    # dt = float(input("input [dt] value: ").replace(",", "."))

    i = ImplicitDiffScheme(h, sigma, L, dt, Nt)
    i.solution()

    e = ExplicitDiffScheme(h, sigma, L, dt, Nt)
    e.solution()

    # k = KrankNicholsonScheme(h, sigma, L, dt, Nt)
    # k.solution()
    #
    # s = SaulevaDiffScheme(h, sigma, L, dt, Nt)
    # s.solution()

    del i
    del e
    # del k
    # del s
